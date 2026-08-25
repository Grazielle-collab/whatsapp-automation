from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
import re
from datetime import datetime

class LeadSpreadsheet:
    def __init__(self, filename='data/leads_crm.xlsx'):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        
        # Garante que a pasta data existe
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        self._initialize()
        
    def _initialize(self):
        """Inicializa a planilha com cabeçalhos formatados"""
        # Se o arquivo existe, tenta carregar
        if os.path.exists(self.filename):
            try:
                self.workbook = load_workbook(self.filename)
                self.sheet = self.workbook.active
                print(f"📂 Planilha carregada: {self.filename}")
                return
            except Exception as e:
                print(f"⚠️ Erro ao carregar planilha: {e}")
                print("📝 Criando nova planilha...")
                # Se não conseguir carregar, cria uma nova
                os.remove(self.filename)  # Remove o arquivo corrompido
        
        # Cria nova planilha
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Leads"
        
        # Define cabeçalhos
        headers = [
            'Data Início', 
            'Nome Cliente', 
            'Telefone', 
            'Origem', 
            'Status', 
            'Data Contato', 
            'Interagiu',
            'Mensagem Original',
            'Observações'
        ]
        self.sheet.append(headers)
        
        # Formata cabeçalhos
        self._format_header()
        self.workbook.save(self.filename)
        print(f"✅ Nova planilha criada: {self.filename}")
        
    def _format_header(self):
        """Formata os cabeçalhos da planilha"""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in self.sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
        # Ajusta largura das colunas
        column_widths = {
            'A': 20,  # Data
            'B': 30,  # Nome
            'C': 18,  # Telefone
            'D': 15,  # Origem
            'E': 12,  # Status
            'F': 20,  # Data Contato
            'G': 12,  # Interagiu
            'H': 50,  # Mensagem
            'I': 30   # Observações
        }
        
        for col, width in column_widths.items():
            self.sheet.column_dimensions[col].width = width
            
    def add_lead(self, lead_data):
        """Adiciona um novo lead à planilha"""
        # Verifica se já existe (evita duplicatas por telefone)
        if self._lead_exists(lead_data.get('telefone', '')):
            print(f"⚠️ Lead já existe: {lead_data.get('nome', 'Desconhecido')}")
            return False
            
        row = [
            lead_data.get('data', datetime.now().strftime('%Y-%m-%d %H:%M')),
            lead_data.get('nome', ''),
            lead_data.get('telefone', ''),
            'Nagato Bot',  # Origem fixa para este caso
            'Novo Lead',
            '',  # Data de contato (será preenchida manualmente)
            'Não',
            lead_data.get('mensagem_original', '')[:200],
            ''  # Observações
        ]
        
        self.sheet.append(row)
        self.workbook.save(self.filename)
        
        # Formata a nova linha
        self._format_new_row(self.sheet.max_row)
        
        print(f"✅ Lead adicionado: {lead_data.get('nome', 'Desconhecido')} - {lead_data.get('telefone', '')}")
        return True
        
    def add_batch_leads(self, leads):
        """Adiciona múltiplos leads de uma vez"""
        added_count = 0
        
        for lead in leads:
            if self.add_lead(lead):
                added_count += 1
                
        print(f"\n📊 Resumo: {added_count} novos leads adicionados de {len(leads)} encontrados")
        return added_count
        
    def _lead_exists(self, telefone):
        """Verifica se o telefone já existe na planilha"""
        if not telefone:
            return False
            
        # Remove formatação para comparação
        clean_phone = re.sub(r'[^\d]', '', str(telefone))
        
        # Itera sobre as linhas (pula o cabeçalho)
        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            cell_phone = row[2].value
            if cell_phone:
                clean_cell = re.sub(r'[^\d]', '', str(cell_phone))
                if clean_phone == clean_cell:
                    return True
        return False
        
    def _format_new_row(self, row_num):
        """Formata a nova linha adicionada"""
        # Destaca novos leads com cor verde claro
        green_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
        
        for col in range(1, 10):
            cell = self.sheet.cell(row=row_num, column=col)
            cell.fill = green_fill
            
    def mark_as_interacted(self, telefone):
        """Marca um lead como interagido"""
        clean_target = re.sub(r'[^\d]', '', str(telefone))
        
        for row in self.sheet.iter_rows(min_row=2, max_col=7):
            cell_phone = row[2].value
            if cell_phone:
                clean_cell = re.sub(r'[^\d]', '', str(cell_phone))
                if clean_cell == clean_target:
                    row[6].value = 'Sim'  # Interagiu
                    row[5].value = datetime.now().strftime('%Y-%m-%d %H:%M')  # Data Contato
                    self.workbook.save(self.filename)
                    print(f"✅ Lead {row[1].value} marcado como interagido")
                    return True
        return False