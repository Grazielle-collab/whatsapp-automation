from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

def create_spreadsheet():
    """Cria uma nova planilha do zero"""
    
    # Garante que a pasta existe
    os.makedirs('data', exist_ok=True)
    
    filename = 'data/leads_crm.xlsx'
    
    # Se existe, deleta
    if os.path.exists(filename):
        os.remove(filename)
        print(f"🗑️ Arquivo antigo removido: {filename}")
    
    # Cria nova planilha
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Leads"
    
    # Cabeçalhos
    headers = [
        'Data Início', 
        'Nome Cliente', 
        'Telefone', 
        'Origem', 
        'Status', 
        'Data Contato', 
        'Interagiu',
        'Mensagem Original',
        'Observações'
    ]
    sheet.append(headers)
    
    # Formata cabeçalhos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Ajusta largura das colunas
    column_widths = {
        'A': 20,  # Data
        'B': 30,  # Nome
        'C': 18,  # Telefone
        'D': 15,  # Origem
        'E': 12,  # Status
        'F': 20,  # Data Contato
        'G': 12,  # Interagiu
        'H': 50,  # Mensagem
        'I': 30   # Observações
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Salva
    wb.save(filename)
    print(f"✅ Planilha criada com sucesso: {filename}")
    print(f"📁 Caminho completo: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    create_spreadsheet()